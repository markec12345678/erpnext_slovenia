"""
Računovodske integracije - izvoz v slovenske računovodske programe.

Podpira izvoz računov in transakcij v:
1. Pantheon (Datex) - najbolj razširjeni slovenski računovodski program
2. e-Račun SI (elektronski arhiv računov)
3. MiniMAX računovodstvo
4. Sigma Controlling

Formati:
- Pantheon: CSV z njihovo strukturo (artikli, stranke, dokumenti)
- e-Račun SI: e-Slog 1.6 XML (že implementirano v einvoice modulu)
- MiniMAX: CSV z_MB Partner kartica formatom
- Sigma Controlling: Excel (XLSX) z več listi

Uporaba:
- RPC: export_to_pantheon(), export_to_minimax(), export_to_sigma()
- Hook: ob submitu Sales/Purchase Invoice generira batch CSV
"""
import csv
import io
import datetime
import frappe
from frappe import _
from frappe.utils.file_manager import save_file


# === PANTHEON EXPORT ===
# Pantheon (Datex) CSV format - Slovenian standard accounting software
# Format: document header + line items, one row per line item

def export_to_pantheon(from_date: str, to_date: str, company: str = None,
                      doctype: str = "Sales Invoice") -> dict:
    """Export Sales/Purchase Invoices to Pantheon CSV format.

    Args:
        from_date: Start date (YYYY-MM-DD)
        to_date: End date (YYYY-MM-DD)
        company: Company name (optional, uses all if not specified)
        doctype: "Sales Invoice" or "Purchase Invoice"

    Returns:
        Dict with file URL and metadata.
    """
    if doctype not in ("Sales Invoice", "Purchase Invoice"):
        frappe.throw(_("Doctype mora biti Sales Invoice ali Purchase Invoice."))

    filters = {
        "docstatus": 1,
        "posting_date": ["between", [from_date, to_date]],
    }
    if company:
        filters["company"] = company

    invoices = frappe.get_all(doctype, filters=filters,
        fields=["name", "posting_date", "customer" if doctype == "Sales Invoice" else "supplier",
                "customer_name" if doctype == "Sales Invoice" else "supplier_name",
                "net_total", "total_taxes_and_charges", "grand_total"])

    if not invoices:
        return {"success": False, "error": _("Ni najdenih računov v izbranem obdobju.")}

    # Build CSV
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)

    # Pantheon header
    writer.writerow([
        "Tip dokumenta", "Številka", "Datum", "Partner - Šifra", "Partner - Naziv",
        "Partner - Davčna št.", "Konto", "Opis", "Znesek brez DDV", "Znesek DDV",
        "Skupni znesek", "Valuta", "Datum zapadlosti", "Sklic SI",
    ])

    for inv in invoices:
        party_field = "customer" if doctype == "Sales Invoice" else "supplier"
        party_name_field = "customer_name" if doctype == "Sales Invoice" else "supplier_name"
        party = inv[party_field]
        party_name = inv[party_name_field]

        # Get party tax_id
        tax_id = frappe.db.get_value(doctype.replace(" Invoice", ""), party, "tax_id") or ""

        # Get due date
        due_date = frappe.db.get_value(doctype, inv.name, "due_date") or inv.posting_date

        # Generate SI reference (MOD 11)
        si_ref = _generate_si_reference(inv.name)

        # Get items
        items = frappe.get_all(f"{doctype} Item",
            filters={"parent": inv.name},
            fields=["item_code", "item_name", "qty", "rate", "net_amount", "income_account"])

        if not items:
            continue

        # Write line per item
        for item in items:
            konto = item.income_account or "400"  # default revenue account
            writer.writerow([
                "Račun" if doctype == "Sales Invoice" else "Prejeti račun",
                inv.name,
                inv.posting_date,
                party[:30] if party else "",
                party_name[:100] if party_name else "",
                tax_id,
                konto,
                item.item_name[:200] if item.item_name else "",
                f"{item.net_amount:.2f}".replace(".", ","),
                "",  # tax per line - computed at document level
                f"{inv.grand_total:.2f}".replace(".", ","),
                "EUR",
                due_date,
                si_ref,
            ])

    csv_content = output.getvalue()
    output.close()

    # Save as file
    file_name = f"Pantheon_export_{doctype.replace(' ', '_')}_{from_date}_{to_date}.csv"
    f = save_file(
        fname=file_name,
        content=csv_content.encode("utf-8-sig"),  # UTF-8 with BOM for Excel
        dt="Company",
        dn=company or frappe.db.get_single_value("Global Defaults", "default_company"),
        folder=None,
        is_private=0,
        df=None,
    )
    frappe.db.commit()

    return {
        "success": True,
        "file_url": f.file_url,
        "file_name": f.file_name,
        "file_size": f.file_size,
        "invoice_count": len(invoices),
        "doctype": doctype,
        "from_date": from_date,
        "to_date": to_date,
    }


# === MINIMAX EXPORT ===
# MiniMAX format - simpler CSV for partner card reconciliation

def export_to_minimax(from_date: str, to_date: str, company: str = None) -> dict:
    """Export Sales/Purchase Invoices to MiniMAX CSV format (Partner Card).

    Returns dict with file URL.
    """
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)

    writer.writerow([
        "Datum", "Številka dokumenta", "Vrsta", "Sifra partnerja", "Naziv partnerja",
        "Davčna številka", "Datum zapadlosti", "Znesek brez DDV", "DDV 22%",
        "DDV 8,5%", "DDV 5%", "Skupni znesek", "Sklic SI", "Opomba",
    ])

    for doctype, doc_type_short in [("Sales Invoice", "Izhodni"), ("Purchase Invoice", "Vhodni")]:
        filters = {
            "docstatus": 1,
            "posting_date": ["between", [from_date, to_date]],
        }
        if company:
            filters["company"] = company

        invoices = frappe.get_all(doctype, filters=filters,
            fields=["name", "posting_date", "customer" if doctype == "Sales Invoice" else "supplier",
                    "customer_name" if doctype == "Sales Invoice" else "supplier_name",
                    "net_total", "total_taxes_and_charges", "grand_total"])

        for inv in invoices:
            party_field = "customer" if doctype == "Sales Invoice" else "supplier"
            party_name_field = "customer_name" if doctype == "Sales Invoice" else "supplier_name"
            party = inv[party_field]
            party_name = inv[party_name_field]

            # Get party tax_id
            tax_id = frappe.db.get_value(doctype.replace(" Invoice", ""), party, "tax_id") or ""
            due_date = frappe.db.get_value(doctype, inv.name, "due_date") or inv.posting_date

            # Get tax breakdown
            taxes = frappe.get_all(f"{doctype.split()[0]} Taxes and Charges",
                filters={"parent": inv.name},
                fields=["rate", "tax_amount"])

            tax_22 = sum(t.tax_amount for t in taxes if abs(t.rate - 22.0) < 0.01)
            tax_85 = sum(t.tax_amount for t in taxes if abs(t.rate - 8.5) < 0.01)
            tax_5 = sum(t.tax_amount for t in taxes if abs(t.rate - 5.0) < 0.01)

            si_ref = _generate_si_reference(inv.name)

            writer.writerow([
                inv.posting_date,
                inv.name,
                doc_type_short,
                party[:20] if party else "",
                party_name[:100] if party_name else "",
                tax_id,
                due_date,
                f"{inv.net_total:.2f}".replace(".", ","),
                f"{tax_22:.2f}".replace(".", ","),
                f"{tax_85:.2f}".replace(".", ","),
                f"{tax_5:.2f}".replace(".", ","),
                f"{inv.grand_total:.2f}".replace(".", ","),
                si_ref,
                f"{doc_type_short} račun",
            ])

    csv_content = output.getvalue()
    output.close()

    file_name = f"MiniMAX_partner_kartica_{from_date}_{to_date}.csv"
    f = save_file(
        fname=file_name,
        content=csv_content.encode("utf-8-sig"),
        dt="Company",
        dn=company or frappe.db.get_single_value("Global Defaults", "default_company"),
        folder=None,
        is_private=0,
        df=None,
    )
    frappe.db.commit()

    return {
        "success": True,
        "file_url": f.file_url,
        "file_name": f.file_name,
        "file_size": f.file_size,
        "from_date": from_date,
        "to_date": to_date,
    }


# === SIGMA CONTROLLING EXPORT ===
# Sigma Controlling expects Excel with multiple sheets

def export_to_sigma(from_date: str, to_date: str, company: str = None) -> dict:
    """Export accounting data to Sigma Controlling XLSX format.

    Returns dict with file URL. Creates Excel with sheets:
    - Bilanca (Balance Sheet accounts)
    - Poslovni izid (P&L accounts)
    - Stranke (Customers)
    - Dobavitelji (Suppliers)
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        return {"success": False, "error": _("openpyxl knjižnica ni nameščena.")}

    wb = Workbook()

    # Sheet 1: Bilanca (Balance Sheet)
    ws1 = wb.active
    ws1.title = "Bilanca"
    ws1.append(["Konto", "Naziv konta", "Tip", "Saldo (EUR)"])
    accounts = frappe.get_all("Account",
        filters={"company": company, "is_group": 0,
                 "root_type": ["in", ["Asset", "Liability", "Equity"]]},
        fields=["account_number", "account_name", "root_type", "name"])
    for acc in accounts:
        # Compute balance from GL Entry
        balance = frappe.db.sql("""
            SELECT COALESCE(SUM(debit - credit), 0)
            FROM `tabGL Entry`
            WHERE account = %s AND is_cancelled = 0
            AND posting_date BETWEEN %s AND %s
        """, (acc.name, from_date, to_date))[0][0] or 0
        ws1.append([acc.account_number, acc.account_name, acc.root_type, float(balance)])

    # Sheet 2: Poslovni izid (P&L)
    ws2 = wb.create_sheet("Poslovni izid")
    ws2.append(["Konto", "Naziv konta", "Tip", "Saldo (EUR)"])
    accounts = frappe.get_all("Account",
        filters={"company": company, "is_group": 0,
                 "root_type": ["in", ["Income", "Expense"]]},
        fields=["account_number", "account_name", "root_type", "name"])
    for acc in accounts:
        balance = frappe.db.sql("""
            SELECT COALESCE(SUM(debit - credit), 0)
            FROM `tabGL Entry`
            WHERE account = %s AND is_cancelled = 0
            AND posting_date BETWEEN %s AND %s
        """, (acc.name, from_date, to_date))[0][0] or 0
        ws2.append([acc.account_number, acc.account_name, acc.root_type, float(balance)])

    # Sheet 3: Stranke (Customers)
    ws3 = wb.create_sheet("Stranke")
    ws3.append(["Šifra", "Naziv", "Davčna št.", "Promet (EUR)", "Št. računov"])
    customers = frappe.get_all("Customer", fields=["name", "customer_name", "tax_id"])
    for c in customers:
        sales = frappe.db.sql("""
            SELECT COUNT(*), COALESCE(SUM(grand_total), 0)
            FROM `tabSales Invoice`
            WHERE customer = %s AND docstatus = 1
            AND posting_date BETWEEN %s AND %s
        """, (c.name, from_date, to_date))[0]
        if sales[0] > 0:
            ws3.append([c.name, c.customer_name, c.tax_id or "", float(sales[1]), int(sales[0])])

    # Sheet 4: Dobavitelji (Suppliers)
    ws4 = wb.create_sheet("Dobavitelji")
    ws4.append(["Šifra", "Naziv", "Davčna št.", "Promet (EUR)", "Št. računov"])
    suppliers = frappe.get_all("Supplier", fields=["name", "supplier_name", "tax_id"])
    for s in suppliers:
        purchases = frappe.db.sql("""
            SELECT COUNT(*), COALESCE(SUM(grand_total), 0)
            FROM `tabPurchase Invoice`
            WHERE supplier = %s AND docstatus = 1
            AND posting_date BETWEEN %s AND %s
        """, (s.name, from_date, to_date))[0]
        if purchases[0] > 0:
            ws4.append([s.name, s.supplier_name, s.tax_id or "", float(purchases[1]), int(purchases[0])])

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f"Sigma_Controlling_{from_date}_{to_date}.xlsx"
    f = save_file(
        fname=file_name,
        content=output.getvalue(),
        dt="Company",
        dn=company or frappe.db.get_single_value("Global Defaults", "default_company"),
        folder=None,
        is_private=0,
        df=None,
    )
    frappe.db.commit()

    return {
        "success": True,
        "file_url": f.file_url,
        "file_name": f.file_name,
        "file_size": f.file_size,
        "sheets": ["Bilanca", "Poslovni izid", "Stranke", "Dobavitelji"],
        "from_date": from_date,
        "to_date": to_date,
    }


# === HELPERS ===

def _generate_si_reference(invoice_name: str) -> str:
    """Generate SI reference with MOD 11 checksum from invoice name."""
    import re
    digits = re.sub(r"\D", "", invoice_name)
    if not digits:
        return ""
    # Pad to at least 5 digits
    digits = digits.zfill(5)[:18]
    weights = [2, 3, 4, 5, 6, 7]
    s = 0
    for i, d in enumerate(reversed(digits)):
        s += int(d) * weights[i % 6]
    mod = s % 11
    if mod == 0:
        cs = 0
    elif mod == 1:
        return ""
    else:
        cs = 11 - mod
    return f"SI00{digits}{cs}"


@frappe.whitelist()
def list_supported_programs() -> list:
    """RPC: Return list of supported accounting programs."""
    return [
        {
            "code": "pantheon",
            "name": "Pantheon (Datex)",
            "format": "CSV",
            "description": "Najbolj razširjeni slovenski računovodski program",
            "endpoint": "erpnext_slovenia.accounting_integration.accounting_integration.export_to_pantheon",
        },
        {
            "code": "minimax",
            "name": "MiniMAX",
            "format": "CSV (Partner kartica)",
            "description": "CSV format za partner kartica",
            "endpoint": "erpnext_slovenia.accounting_integration.accounting_integration.export_to_minimax",
        },
        {
            "code": "sigma",
            "name": "Sigma Controlling",
            "format": "XLSX",
            "description": "Excel z več listi (Bilanca, Poslovni izid, Stranke, Dobavitelji)",
            "endpoint": "erpnext_slovenia.accounting_integration.accounting_integration.export_to_sigma",
        },
        {
            "code": "eslog",
            "name": "e-Slog 1.6",
            "format": "XML",
            "description": "Standard za elektronske račune v Sloveniji (že implementirano)",
            "endpoint": "erpnext_slovenia.einvoice.einvoice.get_eslog_xml",
        },
    ]
